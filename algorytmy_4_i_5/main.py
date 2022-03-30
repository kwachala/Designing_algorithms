import string
import random
from openpyxl import Workbook, load_workbook


class robot:
    def __init__(self, identyfikator, typ, masa, zasieg, rozdzielczosc):
        self.identyfikator = identyfikator
        self.typ = typ
        self.masa = masa
        self.zasieg = zasieg
        self.rozdzielczosc = rozdzielczosc


def id_generator(size=3, chars=string.ascii_uppercase + string.digits):
    return ''.join(random.choice(chars) for _ in range(size))

typ = ['AUV', 'AFV', 'AGV']
def random_robot():
    typ = ['AUV', 'AFV', 'AGV']
    return id_generator(), random.choice(typ), random.randint(50, 60), random.randint(1, 10), random.randint(1, 30)


def multiple_robots(M):
    vector = []
    for i in range(M):
        rr = random_robot()
        new_robot = robot(rr[0], rr[1], rr[2], rr[3], rr[4])
        vector.append(new_robot)
    return vector


def show_robots(vector):
    for i in range(len(vector)):
        print(vector[i].identyfikator, vector[i].typ, vector[i].masa, vector[i].zasieg, vector[i].rozdzielczosc)


def save_to_file(vector):
    wb = Workbook()
    sheet = wb.active
    for i in range(len(vector)):
        sheet.append(
            (vector[i].identyfikator, vector[i].typ, vector[i].masa, vector[i].zasieg, vector[i].rozdzielczosc))
    wb.save('C:/Users/drago/OneDrive/Pulpit/roboty.xlsx')


def read_file():
    wb = load_workbook('C:/Users/drago/OneDrive/Pulpit/roboty.xlsx')
    sheet = wb.active
    for row in sheet.iter_rows():
        for cell in row:
            print(cell.value, end=" ")
        print()

#vec = multiple_robots(5)
#show_robots(vec)
#save_to_file(vec)